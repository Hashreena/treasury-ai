"""
Treasury AI — Document Extraction
---------------------------------
Turns uploaded files (PDF, images, Excel, CSV) into structured
Invoice / BankTxn records.

Reality check for the hackathon:
  * CSV / Excel extraction below is REAL and works.
  * PDF text extraction is real for digital PDFs (pdfplumber).
  * Scanned images / screenshots need OCR (Tesseract) or a
    vision LLM — stubbed here. For the demo, prefer CSV/XLSX
    sample data; wire OCR only if you have spare time.
"""
from __future__ import annotations
import csv
import io
from .reconciliation import Invoice, BankTxn


# ---------- Public API ---------------------------------------------------
def extract_invoices(filename: str, content: bytes) -> list[Invoice]:
    rows = _read_tabular(filename, content)
    out: list[Invoice] = []
    for r in rows:
        out.append(Invoice(
            id=_get(r, "id", "invoice", "invoice_id", "invoice no"),
            vendor=_get(r, "vendor", "supplier", "payee"),
            amount=_to_float(_get(r, "amount", "total", "value")),
            currency=_get(r, "currency", "ccy").upper() or "USD",
            reference=_get(r, "reference", "ref", "po", "po number"),
            date=_get(r, "date", "invoice date", "issued"),
        ))
    return out


def extract_bank_txns(filename: str, content: bytes) -> list[BankTxn]:
    rows = _read_tabular(filename, content)
    out: list[BankTxn] = []
    for r in rows:
        out.append(BankTxn(
            id=_get(r, "id", "txn", "transaction", "transaction id"),
            description=_get(r, "description", "desc", "narrative", "details"),
            amount=_to_float(_get(r, "amount", "value", "debit", "credit")),
            currency=_get(r, "currency", "ccy").upper() or "USD",
            reference=_get(r, "reference", "ref", "po"),
            date=_get(r, "date", "value date", "posted"),
        ))
    return out


# ---------- Tabular readers (CSV + Excel) --------------------------------
def _read_tabular(filename: str, content: bytes) -> list[dict]:
    name = filename.lower()
    if name.endswith(".csv") or name.endswith(".tsv"):
        return _read_csv(content, delimiter="\t" if name.endswith(".tsv") else ",")
    if name.endswith(".xlsx") or name.endswith(".xls"):
        return _read_excel(content)
    if name.endswith(".pdf"):
        return _read_pdf(content)
    # images / screenshots — OCR stub
    raise ValueError(
        f"Unsupported file type for '{filename}'. "
        f"Use CSV/XLSX, or add OCR for scanned images."
    )


def _read_csv(content: bytes, delimiter: str = ",") -> list[dict]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    return [{(k or "").strip().lower(): (v or "").strip()
             for k, v in row.items()} for row in reader]


def _read_excel(content: bytes) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl not installed — pip install openpyxl") from exc
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    out: list[dict] = []
    for raw in rows[1:]:
        out.append({headers[i]: ("" if v is None else str(v).strip())
                    for i, v in enumerate(raw) if i < len(headers)})
    return out


def _read_pdf(content: bytes) -> list[dict]:
    """Digital-PDF text extraction. Falls back gracefully if pdfplumber
    is missing. For scanned PDFs you would route to OCR instead."""
    try:
        import pdfplumber
    except ImportError as exc:
        raise RuntimeError("pdfplumber not installed — pip install pdfplumber") from exc
    rows: list[dict] = []
    with pdfplumber.open(io.BytesIO(content)) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables() or []:
                if not table or len(table) < 2:
                    continue
                headers = [str(h).strip().lower() if h else "" for h in table[0]]
                for raw in table[1:]:
                    rows.append({headers[i]: (c or "").strip()
                                 for i, c in enumerate(raw) if i < len(headers)})
    if not rows:
        raise ValueError("No tables found in PDF — may be scanned; OCR needed.")
    return rows


# ---------- Helpers ------------------------------------------------------
def _get(row: dict, *keys: str) -> str:
    for k in keys:
        if k in row and row[k]:
            return str(row[k]).strip()
    return ""


def _to_float(value: str) -> float:
    if not value:
        return 0.0
    cleaned = value.replace(",", "").replace("$", "").replace("€", "").strip()
    try:
        return round(float(cleaned), 2)
    except ValueError:
        return 0.0
